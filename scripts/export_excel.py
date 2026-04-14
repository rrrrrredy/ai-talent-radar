#!/usr/bin/env python3
"""
AI人才雷达 - Excel 导出模块
支持将搜索结果导出为 Excel 文件

⚠️ 安全限制：
1. 导出前需用户明确确认
2. 单次最多导出50条记录
3. 结果超过10条时询问用户
"""

import json
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# 导出限制常量
MAX_EXPORT_RECORDS = 50    # 单次最多导出记录数
CONFIRM_THRESHOLD = 10     # 超过此数量需用户确认


def confirm_export(record_count: int) -> bool:
    """
    询问用户是否确认导出
    
    Args:
        record_count: 要导出的记录数量
    
    Returns:
        bool: 用户是否同意导出
    """
    if record_count <= 0:
        print("没有数据需要导出")
        return False
    
    if record_count > MAX_EXPORT_RECORDS:
        print(f"⚠️  数据量过大（{record_count}条），超过单次导出限制（{MAX_EXPORT_RECORDS}条）")
        print(f"   将只导出前 {MAX_EXPORT_RECORDS} 条记录")
        record_count = MAX_EXPORT_RECORDS
    
    if record_count > CONFIRM_THRESHOLD:
        print(f"\n📊 即将导出 {record_count} 条人才数据为 Excel")
        print("   用途声明：仅供当前招聘评估使用")
        print("   数据范围：公开学术和开源信息")
        print("   确认导出？(回复 '确认' 继续)")
        # 实际场景中等待用户输入
        # 这里模拟需要确认
        return False  # 默认需要用户确认
    
    return True


def export_search_results(data: dict, output_path: str, force: bool = False) -> bool:
    """
    导出搜索结果到 Excel
    
    Args:
        data: 搜索数据
        output_path: 输出文件路径
        force: 是否跳过确认（仅供内部使用）
    
    Returns:
        bool: 是否成功导出
    """
    # 统计记录数
    total_records = 0
    total_records += len(data.get("academic_talents", []))
    total_records += len(data.get("engineering_talents", []))
    total_records += len(data.get("aminer_scholars", []))
    
    # 用户确认
    if not force and not confirm_export(total_records):
        print("导出已取消。如需导出，请明确回复 '确认'")
        return False
    
    # 限制导出数量
    if total_records > MAX_EXPORT_RECORDS:
        print(f"⚠️  数据量超过限制，仅导出前 {MAX_EXPORT_RECORDS} 条")
    
    wb = Workbook()
    
    # 删除默认 sheet
    wb.remove(wb.active)
    
    # 创建学术人才 sheet
    academic_talents = data.get("academic_talents", [])[:MAX_EXPORT_RECORDS//3]
    if academic_talents:
        ws_academic = wb.create_sheet("学术人才")
        _write_academic_talents(ws_academic, academic_talents)
    
    # 创建工程人才 sheet
    engineering_talents = data.get("engineering_talents", [])[:MAX_EXPORT_RECORDS//3]
    if engineering_talents:
        ws_eng = wb.create_sheet("工程人才")
        _write_engineering_talents(ws_eng, engineering_talents)
    

    
    # 创建汇总 sheet
    ws_summary = wb.create_sheet("汇总", 0)
    _write_summary(ws_summary, data, total_records)
    
    # 保存
    try:
        wb.save(output_path)
        print(f"✅ Excel 已导出: {output_path}")
        print(f"   记录数: {min(total_records, MAX_EXPORT_RECORDS)} 条")
        print(f"   用途限制: 仅供当前招聘使用")
        return True
    except Exception as e:
        print(f"❌ 导出失败: {e}")
        return False


def _write_summary(ws, data, total_records):
    """写入汇总信息"""
    ws.append(["AI人才雷达 - 搜索报告"])
    ws.append([f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M')}"])
    ws.append([f"搜索关键词: {data.get('query', 'N/A')}"])
    ws.append([f"总记录数: {total_records} (导出 {min(total_records, MAX_EXPORT_RECORDS)} 条)"])
    ws.append([])
    ws.append(["⚠️  使用声明"])
    ws.append(["1. 数据来源: 公开学术数据库和开源平台"])
    ws.append(["2. 使用范围: 仅供当前招聘评估使用"])
    ws.append(["3. 隐私保护: 不包含任何个人隐私数据"])
    ws.append(["4. 禁止行为: 禁止二次分发或用于其他商业目的"])
    ws.append([])
    
    insights = data.get("combined_insights", {})
    ws.append(["统计概览"])
    ws.append(["学术人才", insights.get("total_academic_profiles", 0)])
    ws.append(["工程人才", insights.get("total_engineering_profiles", 0)])
    ws.append(["AMiner学者", insights.get("total_aminer_scholars", 0)])
    
    # 设置样式
    ws["A1"].font = Font(size=16, bold=True)
    ws["A6"].font = Font(bold=True, color="FF0000")
    ws["A12"].font = Font(bold=True)
    
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 15


def _write_academic_talents(ws, talents):
    """写入学术人才数据"""
    headers = ["姓名", "机构", "h-index", "引用数", "论文数", "相关论文", "档案链接"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for t in talents:
        papers_str = "; ".join([f"{p['title'][:50]}...({p['year']})" 
                               for p in t.get("relevant_papers", [])[:3]])
        ws.append([
            t.get("name", ""),
            ", ".join(t.get("affiliations", [])) if t.get("affiliations") else "",
            t.get("h_index", 0),
            t.get("citation_count", 0),
            t.get("paper_count", 0),
            papers_str,
            f"https://www.semanticscholar.org/author/{t.get('author_id', '')}"
        ])
    
    # 调整列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["F"].width = 60
    ws.column_dimensions["G"].width = 50


def _write_engineering_talents(ws, talents):
    """写入工程人才数据"""
    headers = ["用户名", "姓名", "简介", "公司", "位置", "Followers", "技术栈", "代表作"]
    ws.append(headers)
    
    # 设置表头样式
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        cell.font = Font(bold=True, color="FFFFFF")
    
    for t in talents:
        featured = t.get("featured_repo", {})
        featured_str = f"{featured.get('name', '')} ({featured.get('stars', 0)}⭐)" if featured else ""
        
        ws.append([
            t.get("username", ""),
            t.get("name", ""),
            t.get("bio", "")[:100] if t.get("bio") else "",
            t.get("company", ""),
            t.get("location", ""),
            t.get("followers", 0),
            ", ".join(t.get("tech_stack", {}).get("top_languages", [])[:5]),
            featured_str
        ])
    
    # 调整列宽
    ws.column_dimensions["A"].width = 15
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 50
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["G"].width = 25
    ws.column_dimensions["H"].width = 40


if __name__ == "__main__":
    import argparse
    
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, help="输入 JSON 文件路径")
    parser.add_argument("--output", required=True, help="输出 Excel 文件路径")
    parser.add_argument("--force", action="store_true", help="跳过确认（仅限内部使用）")
    args = parser.parse_args()
    
    with open(args.input, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    success = export_search_results(data, args.output, force=args.force)
    sys.exit(0 if success else 1)
